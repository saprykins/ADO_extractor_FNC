import datetime
import logging
import azure.functions as func
from azure.storage.blob import BlobServiceClient, BlobClient, ContainerClient
import json
import html2text
import requests
import base64
import pandas as pd
import time
import openpyxl
from io import BytesIO
from openpyxl import Workbook
from openpyxl.utils.dataframe import dataframe_to_rows





# history: https://dev.azure.com/go-gl-pr-migfactory-axa365/Migration_Factory/_apis/wit/workItems/110355/updates?api-version=7.0

pat = 'h*'
organization = 'g*'
project_msft = 'M*'
project_tcs = 'A*'

# blob
connect_str = "D*"
container_name = "c*"

authorization = str(base64.b64encode(bytes(':'+pat, 'ascii')), 'ascii')


# initialization dataFrame
cols_app =  [
    "App id in ADO", 
    "App name", 
    "Environment",
    "State", 
    "Entity",
    "Planned migration date",
    "Actual migration startdate",
    "Actual migration enddate",
    # "Planned Assessment Date", 
    # "Planned Replication Date", 
    # "PlannedPostMigrationDate", 
    # "Planned Design Date", 
    # "Planned Go Live Date",
    "Data center",
    # "Rollback",
    "Blocker details",
    "De-scoping Details",
    "Flow opening confirmation", # not available
    "Last minute reschedule",
    "Migration eligibility",
    "Planned Wave", # not available
    "Internet  access through proxies",
    "Outbound Emails",
    "Reverse Proxies",
    "WAC",
    "WAF",
    "VPN",
    "Load Balancer",
    "Service Account in local AD domains",
    "Encryption",
    "Secret data",
    "Fileshare",
    "Administration through specific Jump servers",
    "Access through specific Citrix Jump servers",
    "Out of business hours",
    "Zero downtime requirements",
    "Risk level",
    "Factory",
    "Sign-off DBA", # NOK
    "Sign-off Entity", # NOK
    # "Last update",
    # "Wave", 
    "Schedule_change_Description"
    ]

cols_servers_msft = ["Server id in ADO", "Server", "FQDN", "Sign-off Ops", "Sign-off DBA"]
cols_servers_tcs = [
    "Server id in ADO", 
    "Server", 
    "FQDN", 
    "Sign-off Ops", 
    "Sign-off DBA",
    "App id in ADO"
    ]

cols_map_servers_apps = ["Server id in ADO", "App id in ADO"]
cols_history =  ["App id in ADO", "Phases"]


df_applications = pd.DataFrame([],  columns = cols_app)
df_servers_msft = pd.DataFrame([],  columns = cols_servers_msft)
df_servers_tcs = pd.DataFrame([],  columns = cols_servers_tcs)
df_map_server_vs_app = pd.DataFrame([],  columns = cols_map_servers_apps)
# df_dates = pd.read_csv('./results/__afa_dates.csv')
# pd.set_option('display.max_rows', df_dates.shape[0]+1)
# print(df_dates)


def get_mig_date(playbook_id):
    try:
        date = df_dates.loc[df_dates["Playbook WI"] == playbook_id, "Mig date"]
    # print(date[0])
    except:
        date = ""
    return date




def get_app_list_for_the_wave_msft(list_of_applications):
    """
    Contains 2 parts: wave2 and entity AFA
    """

    # part 2 (getting microsoft apps)
    url = "https://dev.azure.com/" + organization + "/" + project_msft + "/_apis/wit/wiql/bf60899f-afe1-4701-b5e3-fcd4ae04dd31" # all in ms projects
    # url = "https://dev.azure.com/" + organization + "/" + project_msft + "/_apis/wit/wiql/88efb538-bfa2-4ac0-8c31-1a5d470d5a22" # template only
    
    headers = {
        'Accept': 'application/json',
        'Authorization': 'Basic '+ authorization
    }
    response = requests.get(
        url = url,
        headers=headers,
    )

    list_of_all_ms_applications = [] # 
    applications_raw_data = response.json()["workItems"]
    for application in applications_raw_data:
        list_of_all_ms_applications.append(application["id"])

    
    return list_of_all_ms_applications


def save_application_wi_into_data_frame_msft(application_wi_id, df_applications):   
    """
    Get a working item title, parent, status 
    and saves it into a dataframe
    application_wi_id - the application for which data is extracted
    df_applications - used as storage object
    """
    
    url = 'https://dev.azure.com/' + organization + '/_apis/wit/workItems/' + str(application_wi_id) + '?$expand=all'
    headers = {
        'Accept': 'application/json',
        'Authorization': 'Basic '+ authorization
    }
    response = requests.get(
        url = url,
        headers=headers,
    )

    # list of app attributes
    # is used to use cycles
    app_attributes = []

    # list of keys in ADO
    app_keys_ado = [
        "System.Title", 
        "Custom.EnvironmentTargetSubscription",
        "System.State",
        "Custom.Entity",
        "Custom.PlannedStartDate", # planned cut-over date
        "Custom.MigrationStartDate",
        "Custom.MigrationEndDate", # new field
        # "Custom.PlannedAssessmentDate",
        # "Custom.PlannedReplicationDate",
        
        # The keys below are unavailable in current template of ADO for TCS:
        "Custom.DataCenter",
        "Custom.RollbackReason", # de-scoping or blocker detail -> rollback reason
        "Custom.DeScopingDetails" # should go deeper
        "Custom.DeScopingDetails", # should go deeper
        "Custom.Status2", # FW OK
        "Custom.LastMinuteReschedule", # last minute reschedule
        "Custom.MigrationEligibility", # ok
        "Custom.Wave", # not available
        "Custom.Internetaccessthroughproxies",
        "Custom.OutboundEmails", # ok
        "Custom.ReverseProxies",
        "Custom.WAC",
        "Custom.WAF",
        "Custom.VPN",
        "Custom.LoadBalancer",
        "Custom.ServiceAccountinlocalADdomains",
        "Custom.Encryption",
        "Custom.SecretData",
        "Custom.FileShare",
        "Custom.AdminJumpServer",
        "Custom.AccessthroughspecificCitrixJumpservers",
        "Custom.MigrationConstraint",
        "Custom.ZeroDownTime",
        "Custom.RiskLevel", 
        "Custom.ApplicationOwnershipOrganization",
        "Sign-off DBA",
        "Sign-off Entity",
        # "System.RevisedDate",
        #"Custom.Wave"
        "System.Description"
    ]

    # Try to get data from ADO using keys, 
    # if key not found, save empty space
    for i in range(len(app_keys_ado)):
        try:
            # app_attributes[i+1] = response.json()["fields"][app_keys_ado[i]] # may be need to string
            app_attributes.insert(i+1, response.json()["fields"][app_keys_ado[i]])  # may be need to string
        except: 
            # app_attributes[i+1] = ""
            app_attributes.insert(i+1, "")

    
    # app_attributes[0] = application_wi_id
    app_attributes.insert(0, application_wi_id)
    app_attributes[-4] = "Microsoft"

    # default description 
    default_description_1 = "Add Application all"
    default_description_2 = "Add short description"

    # line with html code that requires text treatment: 
    try:
        # app_attributes[i+1] = response.json()["fields"][app_keys_ado[i]] # may be need to string
        description = response.json()["fields"]["System.Description"]
        description = html2text.html2text(description)
        # print("descr: ", description)
        # message.startswith('Python')
        # if (description.startswith(default_description_1) OR (description.startswith(default_description_2):
        if (((description.startswith(default_description_1)) | (description.startswith(default_description_2)))):
        # if (description.startswith("Add short description")):
            description = ""
            app_attributes[-1] = description
        else:
            app_attributes[-1] = html2text.html2text(description)
    except: 
        # app_attributes[i+1] = ""
        app_attributes[-1] = ""

    # app_attributes.insert(len(app_attributes)+1, "wave_2")
    # add list of servers
    # list_of_ids_of_servers = []
    # list_of_ids_of_servers = get_server_wi_ids_from_application(application_wi_id)

    # new_row = [application_wi_id, wi_title, wi_env, wi_state, wi_entity, wi_date, wi_wave]
    new_row = app_attributes
    new_df = pd.DataFrame([new_row], columns=cols_app)
    
    # load data into a DataFrame object:
    df_applications = pd.concat([df_applications, new_df], ignore_index = True)

    return df_applications


def get_server_wi_ids_from_feature(feature_id):
    """
    Given feature_id the function gets data on its children
    It verified if feature name is "Servers"
    And get its children ids
    """

    url = 'https://dev.azure.com/' + organization + '/_apis/wit/workItems/' + str(feature_id) + '?$expand=all'
    headers = {
        'Accept': 'application/json',
        'Authorization': 'Basic '+ authorization
    }

    response = requests.get(
        url = url,
        headers=headers,
    )
    
    list_of_ids_of_servers = []

    feature_title = response.json()["fields"]["System.Title"]
    if "Servers" in feature_title:
        relations = response.json()["relations"]
        for relation in relations: 
            if relation["rel"] == "System.LinkTypes.Hierarchy-Forward":
                raw_id = relation['url']
                start_line = raw_id.find('workItems/') + 10
                server_id = int(raw_id[start_line:])
                list_of_ids_of_servers.append(server_id)

    return list_of_ids_of_servers



def get_server_wi_ids_from_application(application_id):
    """
    Given app_id, this function gets ids of its servers
    """

    url = 'https://dev.azure.com/' + organization + '/_apis/wit/workItems/' + str(application_id) + '?$expand=all'
    servers_id = []
    headers = {
        'Accept': 'application/json',
        'Authorization': 'Basic '+ authorization
    }
    response = requests.get(
        url = url,
        headers=headers,
    )
    
    # go through features of an app
    # not all applications have servers stored in ADO
    try:
        wi_relations = response.json()["relations"]
    except: 
        wi_relations = ""

    for relation in wi_relations:
        if relation["rel"] == "System.LinkTypes.Hierarchy-Forward":
            # need to go deeper to find servers
            # features can be servers or playbook
            raw_id = relation['url']
            start_line = raw_id.find('workItems/') + 10
            feature_id = int(raw_id[start_line:])
            # print(feature_id) # correct
            list_of_ids_of_servers = get_server_wi_ids_from_feature(feature_id)
            if len(list_of_ids_of_servers)>0:
                # print(list_of_ids_of_servers)
                servers_id = servers_id + list_of_ids_of_servers

        # should we keep it (only 1 feature with servers)
        elif relation["rel"] == "System.LinkTypes.Hierarchy-Reverse":
            # get wave name
            raw_id = relation['url']
            start_line = raw_id.find('workItems/') + 10
            parent_id = int(raw_id[start_line:])
            # print(parent_id)

        # print(relation)
    return servers_id


def save_server_wi_into_data_frame_msft(server_wi_id, df_servers_msft):
    """
    Get a server hostname, statuses
    and saves it into a dataframe
    """
    
    url = 'https://dev.azure.com/' + organization + '/_apis/wit/workItems/' + str(server_wi_id) + '?$expand=all'
    headers = {
        'Accept': 'application/json',
        'Authorization': 'Basic '+ authorization
    }
    response = requests.get(
        url = url,
        headers=headers,
    )
    # server item Title
    wi_title = response.json()["fields"]["System.Title"]

    # server hostname
    try:
        wi_hostname = response.json()["fields"]["Custom.HostName"]
    except: 
        wi_hostname = ""

    # need insert sign-off state    
    sign_off_ops_state = ''
    sign_off_dba_state = ''
    # working item sign-offs DBA
    try:
        sign_off_ops_state = response.json()["fields"]["Custom.SignofffromOpsteam"]
    except: 
        sign_off_ops_state = ""
    
    try:
        sign_off_dba_state = response.json()["fields"]["Custom.SignofffromDBA"]
    except: 
        sign_off_dba_state = ""
    

    new_row = [server_wi_id, wi_title, wi_hostname, sign_off_ops_state, sign_off_dba_state]
    new_df = pd.DataFrame([new_row], columns=cols_servers_msft)

    # load data into a DataFrame object:
    df_servers_msft = pd.concat([df_servers_msft, new_df], ignore_index = True)

    return df_servers_msft


def get_all_servers_list_from_ado_msft():
    """
    The function uses query that is defined in ADO
    The mentioned query displays the list of all servers
    """
    list_of_all_servers = []
    
    url = "https://dev.azure.com/" + organization + "/" + project_msft + "/_apis/wit/wiql/fad91720-c6b5-4e92-be7a-9d98b41d6289" # servers
    # url = "https://dev.azure.com/" + organization + "/" + project_msft + "/_apis/wit/wiql/d3eff7f1-30a7-484a-b0c7-cbe87365dd86" # 2 servers only
    
    headers = {
        'Accept': 'application/json',
        'Authorization': 'Basic '+ authorization
    }
    response = requests.get(
        url = url,
        headers=headers,
    )

    servers_raw_data = response.json()["workItems"]

    for server in servers_raw_data:
        list_of_all_servers.append(server["id"])
    return list_of_all_servers



def get_all_applications_list_from_ado_msft():
    """
    The function uses query that is defined in ADO
    The mentioned query displays the list of all applications (for all waves in the projects)
    The function exists to create mapping between applications and servers
    """
    list_of_all_applications = []
    
    url = "https://dev.azure.com/" + organization + "/" + project_msft + "/_apis/wit/wiql/e2c3101f-d2e2-4156-a57d-53b40a6fec6a"
    # url = "https://dev.azure.com/" + organization + "/" + project_msft + "/_apis/wit/wiql/c144e534-f0a0-436c-894b-81b8db94408a" # only one app
    headers = {
        'Accept': 'application/json',
        'Authorization': 'Basic '+ authorization
    }
    response = requests.get(
        url = url,
        headers=headers,
    )
    applications_raw_data = response.json()["workItems"]
    for application in applications_raw_data:
        list_of_all_applications.append(application["id"])
    return list_of_all_applications







def save_map_server_vs_app(application_wi_id, df_map_server_vs_app): 
    """
    Get a map between ids (servers vs applications)
    """
    
    list_of_servers = get_server_wi_ids_from_application(application_wi_id)
    for server_id_ado in list_of_servers: 
        new_row = [server_id_ado, application_wi_id]
        new_df = pd.DataFrame([new_row], columns=cols_map_servers_apps)
        # load data into a DataFrame object:
        df_map_server_vs_app = pd.concat([df_map_server_vs_app, new_df], ignore_index = True)  
    return df_map_server_vs_app


def save_file_to_storage(file_name, dframe):
    #
    # Saves dataframe to blob as csv
    #
    
    csv_string = dframe.to_csv(index=False)

    # Get a reference to the blob and upload the CSV data
    blob_service_client = BlobServiceClient.from_connection_string(connect_str)
    blob_name = file_name
    blob_client = blob_service_client.get_blob_client(container=container_name, blob=blob_name)
    blob_client.upload_blob(csv_string, overwrite=True)
    # logging.info('uploaded ', file_name)


'''
def save_file_to_excel(file_name, dframe):
    #
    # Saves dataframe to blob as Excel with specified sheet name
    #
    
    # Create a BytesIO object to store Excel data
    #
    #
    excel_string = dframe.to_excel(index=False)
    #
    #
    #
    # Get a reference to the blob and upload the Excel data
    blob_service_client = BlobServiceClient.from_connection_string(connect_str)
    blob_name = file_name
    blob_client = blob_service_client.get_blob_client(container=container_name, blob=blob_name)
    blob_client.upload_blob(excel_string, overwrite=True)
'''

def save_file_to_excel(file_name, dframe):
    #
    # Saves dataframe to blob as Excel with specified sheet name
    #
    
    # Create a BytesIO object to store Excel data
    excel_stream = BytesIO()

    # Create an Excel workbook and add the DataFrame to a sheet
    wb = Workbook()
    ws = wb.active

    for row in dataframe_to_rows(dframe, index=False, header=True):
        ws.append(row)

    # Save the workbook to the BytesIO stream
    wb.save(excel_stream)
    
    # Get the Excel data as bytes from the BytesIO object
    excel_data = excel_stream.getvalue()

    # Get a reference to the blob and upload the Excel data
    blob_service_client = BlobServiceClient.from_connection_string(connect_str)
    blob_name = file_name
    blob_client = blob_service_client.get_blob_client(container=container_name, blob=blob_name)
    blob_client.upload_blob(excel_data, overwrite=True)
    
    
    
# TCS import
def save_application_wi_into_data_frame_tcs(application_wi_id, df_applications):   
    """
    Get a working item title, parent, status 
    and saves it into a dataframe
    application_wi_id - the application for which data is extracted
    df_applications - used as storage object
    """
    
    url = 'https://dev.azure.com/' + organization + '/_apis/wit/workItems/' + str(application_wi_id) + '?$expand=all'
    headers = {
        'Accept': 'application/json',
        'Authorization': 'Basic '+ authorization
    }
    response = requests.get(
        url = url,
        headers=headers,
    )

    # list of app attributes
    # is used to use cycles
    app_attributes = []

    # list of keys in ADO
    app_keys_ado = [
        "System.Title", 
        "Custom.Environment",
        "System.State",
        "Custom.Entity",
        "Custom.202e1741-c1e6-4f30-b29f-d0b52c686578", # planned cut-over date
        "Custom.ActualCutOverDate",
        "Custom.ActualCutOverEndDate", # i konw it doesn't exist
        # "Custom.PlannedAssessmentDate",
        # "Custom.PlannedReplicationDate",
        
        # The keys below are unavailable in current template of ADO for TCS:
        "Custom.DataCenter",
        "Custom.BlockerReason",
        "Custom.DeScopingDetails" # should go deeper
        "Custom.DeScopingDetails", # should go deeper
        "Custom.Status", # FW opening
        "Last minute reschedule", # not available
        "Custom.MigrationEligibility", # ok
        "Custom.Wave",
        "Custom.Internetaccessthroughproxies",
        "Custom.OutboundEmails",
        "Custom.ReverseProxies",
        "WAC", # not available
        "Custom.WAF",
        "Custom.VPN",
        "Custom.LoadBalancer",
        "Custom.ServiceAccountinlocalADdomains", # Service Account in local AD domains
        "Custom.Encryption", 
        "Custom.SecretData",
        "Custom.FileShare",
        "Custom.AdminJumpServer",
        "Custom.AccessthroughspecificCitrixJumpservers",
        "Out of business hours", # not available
        "Zero downtime requirements", # not available
        "Custom.RiskLevel", 
        "Custom.ApplicationOwnershipOrganization",
        "Sign-off DBA", # not available
        "Sign-off Entity", # not available
        # "System.RevisedDate",
        #"Custom.Wave"
        "System.Description"
    ]

    # Try to get data from ADO using keys, 
    # if key not found, save empty space
    for i in range(len(app_keys_ado)):
        try:
            app_attributes.insert(i+1, response.json()["fields"][app_keys_ado[i]])  # may be need to string
        except: 
            # app_attributes[i+1] = ""
            app_attributes.insert(i+1, "")
    

    app_attributes.insert(0, application_wi_id)
    
    # text html line
    try:
        description = response.json()["fields"]["System.Description"]
        app_attributes[-1] = html2text.html2text(description)
    except: 
        app_attributes[-1] = ""

    app_attributes[-4] = "TCS"

    # wi_wave = "wave_2"

    # add list of servers
    list_of_ids_of_servers = []
    # list_of_ids_of_servers = get_server_wi_ids_from_application(application_wi_id)

    new_row = app_attributes

    new_df = pd.DataFrame([new_row], columns=cols_app)
    
    # load data into a DataFrame object:
    df_applications = pd.concat([df_applications, new_df], ignore_index = True)

    return df_applications


def save_server_wi_into_data_frame_tcs(server_wi_id, df_servers_tcs):
    """
    Get a server hostname, statuses
    and saves it into a dataframe
    """
    
    url = 'https://dev.azure.com/' + organization + '/_apis/wit/workItems/' + str(server_wi_id) + '?$expand=all'
    headers = {
        'Accept': 'application/json',
        'Authorization': 'Basic '+ authorization
    }
    response = requests.get(
        url = url,
        headers=headers,
    )
    # server item Title
    wi_title = response.json()["fields"]["System.Title"]

    # server hostname
    try:
        wi_hostname = response.json()["fields"]["Custom.HostName"]
    except: 
        wi_hostname = ""

    try:
        app_wi_id = response.json()["fields"]["System.Parent"]
    except: 
        app_wi_id = ""
    

    
    sign_off_ops_state = ''
    sign_off_dba_state = ''
    # working item sign-offs DBA
    try:
        sign_off_ops_state = response.json()["fields"]["Custom.SignofffromOpsteam"]
    except: 
        sign_off_ops_state = ""
    
    try:
        sign_off_dba_state = response.json()["fields"]["Custom.SignofffromDBA"]
    except: 
        sign_off_dba_state = ""
        
            
    new_row = [server_wi_id, wi_title, wi_hostname, sign_off_ops_state, sign_off_dba_state, app_wi_id]
    new_df = pd.DataFrame([new_row], columns=cols_servers_tcs)

    # load data into a DataFrame object:
    df_servers_tcs = pd.concat([df_servers_tcs, new_df], ignore_index = True)

    return df_servers_tcs


def get_all_servers_list_from_ado_tcs():
    """
    The function uses query that is defined in ADO
    The mentioned query displays the list of all servers
    """
    list_of_all_servers = []
    
    url = "https://dev.azure.com/" + organization + "/" + project_tcs + "/_apis/wit/wiql/5a8fa180-91e7-482c-b7b1-67879234b19a" # all servers
    # url = "https://dev.azure.com/" + organization + "/" + project_tcs + "/_apis/wit/wiql/0c17ef84-b884-40a4-9381-144b3b417a77" # one servers
    headers = {
        'Accept': 'application/json',
        'Authorization': 'Basic '+ authorization
    }
    response = requests.get(
        url = url,
        headers=headers,
    )
    servers_raw_data = response.json()["workItems"]
    for server in servers_raw_data:
        list_of_all_servers.append(server["id"])
    return list_of_all_servers


def get_all_applications_list_from_ado_tcs():
    """
    The function uses query that is defined in ADO
    The mentioned query displays the list of all applications (for all waves in the projects)
    The function exists to create mapping between applications and servers
    """
    list_of_all_applications = []
    
    url = "https://dev.azure.com/" + organization + "/" + project_tcs + "/_apis/wit/wiql/0a894ff4-67d6-4115-b33e-3aa8a5945e3d" # all apps
    # url = "https://dev.azure.com/" + organization + "/" + project_tcs + "/_apis/wit/wiql/412a2a35-95d5-4837-a1af-12d0e2941a20" # one app
    headers = {
        'Accept': 'application/json',
        'Authorization': 'Basic '+ authorization
    }
    response = requests.get(
        url = url,
        headers=headers,
    )
    applications_raw_data = response.json()["workItems"]
    for application in applications_raw_data:
        list_of_all_applications.append(application["id"])
    return list_of_all_applications


#
# History
#

def get_state_changes_msft(application_id, df_history):
    #
    # Generates a csv-file with history of state changes based on application level (MSFT)
    #
    
    url = 'https://dev.azure.com/' + organization + '/' + project_msft + '/_apis/wit/workItems/' + str(application_id) + '/updates?api-version=7.0'
    headers = {
        'Accept': 'application/json',
        'Authorization': 'Basic '+ authorization
    }
    response = requests.get(
        url = url,
        headers=headers,
    )

    # which states app went through
    app_states = []
    app_history = response.json()["value"]
    
    for state_change in app_history: # for each change 
        try:
            state_change_record = state_change["fields"]["System.State"]

            if state_change_record['newValue']: 
                app_states.append(state_change_record['newValue'].lower())

        except:
            assessment_duration = 0

    # display all phases without repetitions:
    representation_of_states_str = ''
 

    if len(app_states) == 1:
        representation_of_states_str = app_states[0]
    elif len(app_states) == 2:
        representation_of_states_str = app_states[0] + ' -> '+ app_states[1]
    else: 
        for i in range(len(app_states)):
            if i == 0:
                representation_of_states_str = app_states[0]
            else:
                representation_of_states_str = representation_of_states_str + ' -> '+ app_states[i]

    new_row = [application_id, representation_of_states_str]
    new_df = pd.DataFrame([new_row], columns=cols_history)
    df_history = pd.concat([df_history, new_df], ignore_index = True)
    return df_history


'''
def get_state_changes_tcs(application_id, df_history):
    #
    # Generates a csv-file with history of state changes based on application level (TCS)
    #
    
    url = 'https://dev.azure.com/' + organization + '/' + project + '/_apis/wit/workItems/' + str(application_id) + '/updates?api-version=7.0'
    headers = {
        'Accept': 'application/json',
        'Authorization': 'Basic '+ authorization
    }
    response = requests.get(
        url = url,
        headers=headers,
    )

    # which states app went through
    app_states = []
    app_history = response.json()["value"]

    
    for state_change in app_history: # for each change 
        try:
            # record is json with 2 objects: old and new values 
            state_change_record = state_change["fields"]["System.State"]
            # print(state_change_record)
            if state_change_record['newValue']: 
                app_states.append(state_change_record['newValue'].lower())

        except:
            assessment_duration = 0



    res = [app_states[0]]
    for i, c in enumerate(app_states[1:]):
        if c != app_states[i]:
            res.append(c)
    
    app_states = res

    # display all phases without repetitions:
    representation_of_states_str = ''


    if len(app_states) == 1:
        representation_of_states_str = app_states[0]
    elif len(app_states) == 2:
        representation_of_states_str = app_states[0] + ' -> '+ app_states[1]
    else: 
        for i in range(len(app_states)):
            if i == 0:
                representation_of_states_str = app_states[0]
            else:
                representation_of_states_str = representation_of_states_str + ' -> '+ app_states[i]

    new_row = [application_id, representation_of_states_str]
    new_df = pd.DataFrame([new_row], columns=cols_history)
    df_history = pd.concat([df_history, new_df], ignore_index = True)
    # print(df_history)
    return df_history
'''


def make_analysis_for_dates_columns():
    # Define your blob storage connection string and container name
    blob_name = "ADO_extract.xlsx"  # The name of the Excel file in the blob

    # Create a BlobServiceClient instance
    blob_service_client = BlobServiceClient.from_connection_string(connect_str)
    blob_client = blob_service_client.get_blob_client(container=container_name, blob=blob_name)

    # Download the blob data as bytes
    excel_data = blob_client.download_blob().readall()

    # Load the Excel data from bytes into a BytesIO object
    excel_stream = BytesIO(excel_data)

    # Load the Excel workbook from the BytesIO object
    wb = openpyxl.load_workbook(excel_stream)
    ws = wb.active

    # Loop through the rows in the worksheet
    for row in ws.iter_rows(min_row=2, values_only=False):
        planned_date = row[9]  # Column J is the 10th column (0-indexed)
        actual_date = row[10]  # Column K is the 11th column (0-indexed)
        actual_date_obj = None
        
        # Check if the actual date is in the future
        if actual_date.value is not None:
            # actual_date_obj = datetime.strptime(actual_date.value, '%Y-%m-%dT%H:%M:%SZ')
            actual_date_obj = datetime.datetime.strptime(actual_date.value, '%Y-%m-%dT%H:%M:%SZ')

            if actual_date_obj > datetime.datetime.now():
                # Replace the planned date with the actual date
                row[9].value = actual_date_obj.strftime('%Y-%m-%dT%H:%M:%SZ')
                row[10].value = None  # Delete the value in column K
                row[11].value = None  # Delete the value in column L
                
    # Create a BlobServiceClient instance
    blob_service_client = BlobServiceClient.from_connection_string(connect_str)
    blob_client = blob_service_client.get_blob_client(container=container_name, blob=blob_name)

    # Save the modified workbook to a BytesIO stream
    excel_stream = BytesIO()
    wb.save(excel_stream)

    # Get the Excel data as bytes from the BytesIO object
    excel_data = excel_stream.getvalue()

    # Upload the modified Excel data to the blob
    blob_client.upload_blob(excel_data, overwrite=True)

    print("Workbook saved to blob successfully.")
    



def main(mytimer: func.TimerRequest) -> None:
    utc_timestamp = datetime.datetime.utcnow().replace(
        tzinfo=datetime.timezone.utc).isoformat()

    if mytimer.past_due:
        logging.info('The timer is past due!')

    #
    # MSFT
    #
    
    # part ms: apps
    list_of_applications = []
    list_of_applications = get_app_list_for_the_wave_msft(list_of_applications)
    
    df_applications = pd.DataFrame([],  columns = cols_app)
    for application_id in list_of_applications: 
        df_applications = save_application_wi_into_data_frame_msft(application_id, df_applications)
    
    save_file_to_storage('__ms_applications_extract.csv', df_applications)
    

    # part ms: servers
    list_of_servers = get_all_servers_list_from_ado_msft()
    df_servers_msft = pd.DataFrame([],  columns = cols_servers_msft)
    for server in list_of_servers:
        df_servers_msft = save_server_wi_into_data_frame_msft(server, df_servers_msft)
    
    save_file_to_storage('__ms_servers_extract.csv', df_servers_msft)
    
    
    
    # part ms: maps of applications with servers
    list_of_all_applications = get_all_applications_list_from_ado_msft()
    df_map_server_vs_app = pd.DataFrame([],  columns = cols_map_servers_apps)
    for application_id in list_of_all_applications: 
        df_map_server_vs_app = save_map_server_vs_app(application_id, df_map_server_vs_app)
    
    save_file_to_storage('__ms_mapping.csv', df_map_server_vs_app)
    
    
    #
    # History MSFT
    #
    
    df_history_msft = pd.DataFrame([],  columns = cols_history)
    for application_id in list_of_all_applications:
        df_history_msft = get_state_changes_msft(application_id, df_history_msft)

    save_file_to_storage('__ms_history.csv', df_history_msft)
    
    
    # Preparing csv-files (MSFT)  
    df3 = pd.merge(df_servers_msft,df_map_server_vs_app, on=['Server id in ADO'])
    df4 = pd.merge(df3,df_applications, on=['App id in ADO'])
    df_4c = df4
    # df_4c = df4.drop(["Unnamed: 0", "Unnamed: 0_x", "Unnamed: 0_y", "Unnamed: 0"], axis=1)
    
    df5 = pd.merge(df_4c, df_history_msft, on=['App id in ADO'])
    
    # df_ms = df5
    # df_ms = df5.drop(["Unnamed: 0", "Unnamed: 0", "App id in ADO"], axis=1)
    df_ms = df5.drop(["App id in ADO"], axis=1)
    
    df4_outer_join = pd.merge(df3,df_applications, on=['App id in ADO'], how = "outer")
    df5_outer_join = df4_outer_join
    # df5_outer_join = df4_outer_join.drop(["Unnamed: 0", "Unnamed: 0_x", "Unnamed: 0_y", "Unnamed: 0"], axis=1)


    save_file_to_excel("ADO_MS_outer_join.xlsx", df5_outer_join)
    save_file_to_excel("ADO_MS_extract.xlsx", df_ms)
    
    
    #
    # TCS
    #
    
    
    
    list_of_applications = []
    list_of_applications = get_all_applications_list_from_ado_tcs()

    # display the table with apps and details
    df_applications = pd.DataFrame([],  columns = cols_app)
    for application_id in list_of_applications: 
        df_applications = save_application_wi_into_data_frame_tcs(application_id, df_applications)

    # print(df_applications)
    save_file_to_storage('__tcs_applications_extract.csv', df_applications)
    

    # get list of servers
    # for each server save into df
    df_servers_tcs = pd.DataFrame([],  columns = cols_servers_tcs)
    list_of_servers = get_all_servers_list_from_ado_tcs()
    for server in list_of_servers:
        df_servers_tcs = save_server_wi_into_data_frame_tcs(server, df_servers_tcs)
    save_file_to_storage('__tcs_servers_extract.csv', df_servers_tcs)
    
    
    # 
    # History TCS
    # 
    # list_of_applications
    # list_of_all_applications = get_all_applications_list_from_ado_tcs()
    df_history_tcs = pd.DataFrame([],  columns = cols_history)
    for application_id in list_of_applications:
        df_history_tcs = get_state_changes_msft(application_id, df_history_tcs)

    save_file_to_storage('__tcs_history.csv', df_history_tcs)
    
    # TCS files
    df3 = pd.merge(df_servers_tcs,df_applications, on=['App id in ADO'], how = "outer")
    df4 = pd.merge(df3, df_history_tcs, on=['App id in ADO'])
    
    df_tcs = df4
    # df_tcs = df4.drop(["Unnamed: 0_x", "Unnamed: 0_y", "App id in ADO"], axis=1)
    df_tcs = df4.drop(["App id in ADO"], axis=1)

    # for internal usage
    df3_outer_join = pd.merge(df_servers_tcs,df_applications, on=['App id in ADO'], how = 'outer')

    df4_outer_join = df3_outer_join
    # df4_outer_join = df3_outer_join.drop(["Unnamed: 0_x", "Unnamed: 0_y"], axis=1)
    
    save_file_to_excel("ADO_TCS_outer_join.xlsx", df4_outer_join)
    # df4_outer_join.to_excel('./results/ADO_TCS_outer_join.xlsx', sheet_name='tcs_all', index=False)
    
    save_file_to_excel("ADO_TCS_extract.xlsx", df_tcs)
    # save_file_to_excel("ADO_TCS_extract.xlsx", df_tcs, "tcs")
    
    
    # UNION
    union_dfs = pd.concat([df_ms, df_tcs])
    # union_dfs = union_dfs.drop(["Unnamed: 0"], axis=1)
    save_file_to_excel("ADO_extract.xlsx", union_dfs)
    # save_file_to_excel("ADO_extract.xlsx", union_dfs, "total")
    
    
    make_analysis_for_dates_columns()
    
    
    logging.info('Python timer trigger function ran at %s', utc_timestamp)
